
# coding: latin1
import os
import pandas as pd
import sys
import pdfplumber
import PyPDF2
import aspose.pdf as ap
import aspose.pdf as pdf
import openpyxl
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import tabula as tb
from tabula import read_pdf
#from pdf2excel import Converter
from openpyxl.styles import PatternFill
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QPushButton, QFileDialog, QMainWindow, QVBoxLayout, QMessageBox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from ast import Lambda
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QGraphicsDropShadowEffect
from PyQt5.QtCore import QPropertyAnimation,QEasingCurve
from PyQt5.QtGui import QColor
from PyQt5.uic import loadUi
from PyQt5.QtCore import pyqtSlot

class rooti(QMainWindow):
    def __init__(self):
        super(rooti,self).__init__()
        loadUi('D:/ESPE/Practicas INEN/interfaz2/PythonApplication1/interfaz2.ui',self)
        self.bt_menu.clicked.connect(self.mover_menu)
        self.bt_restart.hide()
        #condiciones iniciales botones
        self.metodo1.model().item(0).setEnabled(False)
        self.proceso1.model().item(0).setEnabled(False)
        self.metodo1.setEnabled(False)
        self.proceso1.setEnabled(False)
        self.bt_filter1.setEnabled(False)
        self.bt_save1.setEnabled(False)
        self.bt_load1.setEnabled(False)
        self.bt_load2.setEnabled(False)
        self.bt_save2.setEnabled(False)
        self.bt_filter2.setEnabled(False)
        #botones de funciones
        self.bt_equi1.clicked.connect(self.equipo1)
        self.bt_equi2.clicked.connect(self.equipo2)
        self.bt_load1.clicked.connect(self.load1)
        self.bt_load2.clicked.connect(self.load2)
        self.bt_filter1.clicked.connect(self.filter1)
        self.bt_filter2.clicked.connect(self.filter2)
        self.bt_save1.clicked.connect(self.save1)
        self.bt_save2.clicked.connect(self.save2)
        #botones de barra de titulo
        self.bt_maxi.clicked.connect(self.maximizar)
        self.bt_restart.clicked.connect(self.normal)
        self.bt_mini.clicked.connect(self.minimizar)
        self.bt_close.clicked.connect(lambda: self.close())
        
        #eliminar la ventana del main
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setWindowOpacity(1)
        #SizeGrip
        self.gripSize=10
        self.grip=QtWidgets.QSizeGrip(self)
        self.grip.resize(self.gripSize,self.gripSize)
        #mover ventana
        self.frame_top.mouseMoveEvent=self.mover_ventana
        #conexion botones
        self.bt_equi1.clicked.connect(lambda:self.stackedWidget.setCurrentWidget(self.page_equipo1))
        self.bt_equi2.clicked.connect(lambda:self.stackedWidget.setCurrentWidget(self.page_equipo2))
        self.metodo1.currentIndexChanged.connect(self.actualizarComboBoxProceso)

    @pyqtSlot()
    def actualizarComboBoxProceso(self):
            metodo_seleccionado = self.metodo1.currentText()
            self.proceso1.clear()  # Limpiar las opciones actuales en comboBoxProceso1
            self.metodo=metodo_seleccionado
            if metodo_seleccionado == "Método Llama":
                self.proceso1.addItems(["Seleccioné una opción","Opción A"])
                self.proceso1.model().item(0).setEnabled(False)
                
            elif metodo_seleccionado == "Método Horno de Grafito":
                self.proceso1.addItems(["Seleccioné una opción","Opción B"])
                self.proceso1.model().item(0).setEnabled(False)
                
            elif metodo_seleccionado == "Método Generación de Hidruros":
                self.proceso1.addItems(["Seleccioné una opción","Opción C"])
                self.proceso1.model().item(0).setEnabled(False)
                
        
        

    

    def minimizar(self):
        self.showMinimized()
        
    def normal(self):
        self.showNormal()
        self.bt_restart.hide()
        self.bt_maxi.show()
    
    def maximizar(self): 
        self.showMaximized()
        self.bt_restart.show()
        self.bt_maxi.hide()
        
    def mousePressEvent(self,event):
        self.click_position=event.globalPos()
    
    def mover_ventana(self,event):
        if self.isMaximized()==False:
            if event.buttons()==QtCore.Qt.LeftButton:
                self.move(self.pos()+event.globalPos()-self.click_position)
                self.click_position=event.globalPos()
                event.accept()
        if event.globalPos().y()<=10:
            self.showMaximized()
            self.bt_maxi.hide()
            self.bt_restart.show()
        else:
            self.showNormal()
            self.bt_maxi.show()
            self.bt_restart.hide()
                           

 
    def mover_menu(self):
        if True:
            width=self.frame_control.width()
            normal=0
            if width==0:
                extender=300
            else:
                extender=normal
            self.animacion=QPropertyAnimation(self.frame_control,b'minimumWidth')
            self.animacion.setDuration(300)
            self.animacion.setStartValue(width)
            self.animacion.setEndValue(extender)
            self.animacion.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
            self.animacion.start()
    
    def equipo1(self):
        self.bt_load1.setEnabled(True)
        
        # Conectar la señal currentIndexChanged de comboBoxMetodo1 a la función actualizarComboBoxProceso

        
    def load1(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        archivo, _ = QFileDialog.getOpenFileName(self, 'Buscador de csv', '', 'Archivos CSV (*.csv);;Todos los archivos (*)', options=options)
        if archivo:
            self.label_4.setText("Archivo cargado")
            self.archivo_seleccionado = archivo
            self.bt_filter1.setEnabled(True)
            self.metodo1.setEnabled(True)
            self.proceso1.setEnabled(True)
            
            
                       
            
            
            
        else:
            self.label_4.setText('No se seleccionó ningún archivo.')
            self.bt_filter1.setEnabled(False)
    
    

    
    
    
    def filter1(self):
        if self.archivo_seleccionado:
            try: 
                
                metho1=self.metodo1.currentText()
                process1=self.proceso1.currentText()
                df = pd.read_csv(self.archivo_seleccionado, encoding='latin-1')
                columnas = df.columns
                print(columnas)
                
                if metho1=='Método Llama'and process1=='Opción A':
                    tabla = df[['RSD (Corr Abs)','Conc (Calib)','Conc (Samp)','RSD (Conc)','Abs (Corr)1','Conc (Calib)1','Conc (Samp)1', 'Abs (Corr)2','Conc (Calib)2','Conc (Samp)2', 'Abs (Corr)3','Conc (Calib)3','Conc (Samp)3']]
                    self.tabla = tabla
                elif metho1=='Método Horno de Grafito'and process1=='Opción B':
                    tabla = df[['RSD (Corr Abs)','Conc (Calib)','Conc (Samp)','RSD (Conc)','Abs (Corr)1','Conc (Calib)1','Conc (Samp)1', 'Abs (Corr)2','Conc (Calib)2','Conc (Samp)2', 'Abs (Corr)3','Conc (Calib)3','Conc (Samp)3']]
                    self.tabla = tabla
                elif metho1=='Método Generación de Hidruros'and process1=='Opción C':
                    tabla = df[['RSD (Corr Abs)','Conc (Calib)','Conc (Samp)','RSD (Conc)','Abs (Corr)1','Conc (Calib)1','Conc (Samp)1', 'Abs (Corr)2','Conc (Calib)2','Conc (Samp)2', 'Abs (Corr)3','Conc (Calib)3','Conc (Samp)3']]
                    self.tabla = tabla
                
                self.label_7.setText("Has selecionado"+metho1+" y "+process1)    
                self.label_4.setText("Archivo Filtrado")
                self.bt_save1.setEnabled(True) 
                    
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al filtrar la tabla: {str(e)}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se seleccionó ningún archivo.')
    
    
    def save1(self):
        if hasattr(self, 'tabla'):
            try:
                nombre_archivo = os.path.splitext(os.path.basename(self.archivo_seleccionado))[0]
                nuevo_nombre_excel = f"Filtrado_{nombre_archivo}.xlsx"
                options = QFileDialog.Options()
                options |= QFileDialog.ReadOnly
                ruta_guardado, _ = QFileDialog.getSaveFileName(self, 'Guardar en Excel', nuevo_nombre_excel, 'Archivos Excel (*.xls);;Todos los archivos (*)', options=options)
                
                if ruta_guardado:
                    df = pd.DataFrame(self.tabla)
                    writer = pd.ExcelWriter(ruta_guardado, engine='xlsxwriter')
                    df.to_excel(writer, index=False, sheet_name='Hoja1')
                    workbook  = writer.book
                    worksheet = writer.sheets['Hoja1']
                    header_format = workbook.add_format({'bg_color': '#00FF00', 'align': 'center', 'valign': 'vcenter','bold': True,'border':1})
                    for col_num, value in enumerate(df.columns.values):
                        worksheet.write(0, col_num, value, header_format)
                    cell_format=workbook.add_format({ 'align': 'center', 'valign': 'vcenter','border':1})
                    
                    for i, col in enumerate(df.columns):
                        max_len = max(df[col].astype(str).str.len())
                        max_len = max(max_len, len(col))
                        worksheet.set_column(i, i, max_len + 2)
                    worksheet.set_default_row(15)  # Altura de fila predeterminada
                    worksheet.conditional_format(1, 0, len(df), len(df.columns)-1, {'type': 'no_blanks', 'format': cell_format})
                    writer.close()
                    QMessageBox.information(self, 'Información', f'Archivo Excel guardado en: {ruta_guardado}')
                    self.label_4.setText("Archivo Guardado")
                    self.bt_save1.setEnabled(False)
                    self.bt_filter1.setEnabled(False)
                    self.label_7.setText('')
                    self.metodo1.setCurrentIndex(0)
                    self.proceso1.setCurrentIndex(0)
                   
                    
                else:
                    QMessageBox.warning(self, 'Advertencia', 'No se seleccionó una ruta de guardado.')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al guardar en Excel: {str(e)}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se seleccionó ninguna tabla.')
            
   

    def equipo2(self):
        self.bt_load2.setEnabled(True)

    def load2(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(self, "Select PDF File", "", "PDF files (*.pdf)")
        self.archivo_seleccionado = file_path
        if file_path:
            self.label_6.setText("Archivo cargado")
            dataframes = []
            def extract_columns_from_table(table):
                num_columns = len(table[0])
                for col_idx in range(num_columns):
                    column_name = table[0][col_idx]
                    column = [row[col_idx] for row in table[1:]]
                    yield pd.Series(column,name=column_name)
            all_series = []
            with pdfplumber.open(file_path) as pdf:
                 for page in pdf.pages:
       
                    tables = page.extract_tables()
                    for table in tables:
                        for series in extract_columns_from_table(table):
                            all_series.append(series)
                            print(series)
                            print() 
                        
                        print("---- Fin de la tabla ----")
            df = pd.concat(all_series, axis=1)
            print(df)
            self.df = df
            columnas = df.columns
            self.bt_filter2.setEnabled(True)
            print(columnas)
            
            
            
                      

        
            
    def filter2(self):
        if hasattr(self, 'df'):
            self.bt_save2.setEnabled(True)
            
            try:
                
                 df = self.df
                 self.label_6.setText("Archivo Filtrado")
                 
                 df[['Triplicados 1', 'Triplicados 2', 'Triplicados 3']] = df['Triplicados'].str.split(expand=True)
                 print(df[['Triplicados 1', 'Triplicados 2', 'Triplicados 3']])
                 df[['Abs 690 Triplicados 1', 'Abs 690 Triplicados 2', 'Abs 690 Triplicados 3']] = df['Abs 690 Triplicados'].str.split(expand=True)
                 print(df[['Abs 690 Triplicados 1', 'Abs 690 Triplicados 2', 'Abs 690 Triplicados 3']])
                 
                 tabla1 = df[['ID de muestra', 'Media Fosforo Total (mg/L)','Desv est','Media Abs 690 (AU)','Abs 690 Desv est']]
                 tabla1 = pd.concat([tabla1, df[['Triplicados 1', 'Triplicados 2', 'Triplicados 3']],df[['Abs 690 Triplicados 1', 'Abs 690 Triplicados 2', 'Abs 690 Triplicados 3']]], axis=1)
                 print(tabla1)
                 self.tabla1 = tabla1

                
            except Exception as e:
                 QMessageBox.critical(self, 'Error', f'Error al filtrar la tabla: {str(e)}')
            
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se selecci\u00F3n ninguna tabla.')

    def save2(self):
        if hasattr(self, 'tabla1'):
            try:
                nombre_archivo = os.path.splitext(os.path.basename(self.archivo_seleccionado))[0]
                nuevo_nombre_excel = f"Filtrado_{nombre_archivo}.xlsx"
                options = QFileDialog.Options()
                options |= QFileDialog.ReadOnly
                ruta_guardado, _ = QFileDialog.getSaveFileName(self, 'Guardar en Excel', nuevo_nombre_excel, 'Archivos Excel (*.xlsx);;Todos los archivos (*)', options=options)
            
                if ruta_guardado:
                # Crear un DataFrame a partir de self.tabla1 (asegúrate de tener pandas instalado)
                    df = pd.DataFrame(self.tabla1)
                
                # Crear un escritor de Excel
                    writer = pd.ExcelWriter(ruta_guardado, engine='xlsxwriter')
                    df.to_excel(writer, index=False, sheet_name='Hoja1')
                
                # Obtener la hoja de cálculo
                    workbook = writer.book
                    worksheet = workbook.get_worksheet_by_name('Hoja1')
                
                # Formatear la fila de títulos en verde y centrar los datos
                    header_format = workbook.add_format({'bg_color': '#00FF00', 'align': 'center', 'valign': 'vcenter','bold': True,'border':1})
                    for col_num, value in enumerate(df.columns.values):
                        worksheet.write(0, col_num, value, header_format)
                    #header_alignment = Alignment(horizontal="center", vertical="center")

                    cell_format=workbook.add_format({ 'align': 'center', 'valign': 'vcenter','border':1})
                # Ajustar automáticamente el ancho de las columnas
                    for i, col in enumerate(df.columns):
                        max_len = max(df[col].astype(str).str.len())
                        max_len = max(max_len, len(col))
                        worksheet.set_column(i, i, max_len + 2)
                    worksheet.set_default_row(15)  # Altura de fila predeterminada
                    worksheet.conditional_format(1, 0, len(df), len(df.columns)-1, {'type': 'no_blanks', 'format': cell_format})
                # Guardar el archivo Excel
                    writer.close()
                
                    QMessageBox.information(self, 'Informaci\u00F3n', f'Archivo Excel guardado en: {ruta_guardado}')
                    self.label_6.setText("Archivo Guardado")
                    self.bt_save2.setEnabled(False)
                    self.bt_filter2.setEnabled(False)
                else:
                    QMessageBox.warning(self, 'Advertencia', 'No se seleccion\u00F3 una ruta de guardado.')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al guardar en Excel: {str(e)}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se seleccion\u00F3 ninguna tabla.')




    


    
    
            
            
          

        
if __name__=="__main__":
    app=QApplication(sys.argv)
    mi_app=rooti()
    mi_app.show()
    sys.exit(app.exec_())