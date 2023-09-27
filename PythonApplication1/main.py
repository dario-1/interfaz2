# -*- coding: utf-8 -*-
import os
import pandas as pd
import sys
import aspose.pdf as ap
import aspose.pdf as pdf
import openpyxl
import tabula as tb
from tabula import read_pdf
#from pdf2excel import Converter
from openpyxl.styles import PatternFill
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QPushButton, QFileDialog, QMainWindow, QVBoxLayout, QMessageBox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import NamedStyle
from ast import Lambda
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QGraphicsDropShadowEffect
from PyQt5.QtCore import QPropertyAnimation,QEasingCurve
from PyQt5.QtGui import QColor
from PyQt5.uic import loadUi


class rooti(QMainWindow):
    def __init__(self):
        super(rooti,self).__init__()
        loadUi('D:/ESPE/Practicas INEN/interfaz2/PythonApplication1/interfaz2.ui',self)
        self.bt_menu.clicked.connect(self.mover_menu)
        self.bt_restart.hide()
        #condiciones iniciales botones
        self.metodo1.model().item(0).setEnabled(False)
        self.proceso1.model().item(0).setEnabled(False)
        self.bt_filter1.setEnabled(False)
        self.bt_save1.setEnabled(False)
        self.bt_load1.setEnabled(False)
        self.bt_load2.setEnabled(False)
        self.bt_save2.setEnabled(False)
        self.bt_filter2.setEnabled(False)
        #botones de funciones
        self.bt_equi1.clicked.connect(self.equipo1)
        self.bt_equi2.clicked.connect(self.equipo2)
        self.bt_load1.clicked.connect(self.load1)
        self.bt_load2.clicked.connect(self.load2)
        self.bt_filter1.clicked.connect(self.filter1)
        self.bt_filter2.clicked.connect(self.filter2)
        self.bt_save1.clicked.connect(self.save1)
        self.bt_save2.clicked.connect(self.save2)
        #botones de barra de titulo
        self.bt_maxi.clicked.connect(self.maximizar)
        self.bt_restart.clicked.connect(self.normal)
        self.bt_mini.clicked.connect(self.minimizar)
        self.bt_close.clicked.connect(lambda: self.close())
        
        #eliminar la ventana del main
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setWindowOpacity(1)
        #SizeGrip
        self.gripSize=10
        self.grip=QtWidgets.QSizeGrip(self)
        self.grip.resize(self.gripSize,self.gripSize)
        #mover ventana
        self.frame_top.mouseMoveEvent=self.mover_ventana
        #conexion botones
        self.bt_equi1.clicked.connect(lambda:self.stackedWidget.setCurrentWidget(self.page_equipo1))
        self.bt_equi2.clicked.connect(lambda:self.stackedWidget.setCurrentWidget(self.page_equipo2))
        
        

    

    def minimizar(self):
        self.showMinimized()
        
    def normal(self):
        self.showNormal()
        self.bt_restart.hide()
        self.bt_maxi.show()
    
    def maximizar(self): 
        self.showMaximized()
        self.bt_restart.show()
        self.bt_maxi.hide()
        
    def mousePressEvent(self,event):
        self.click_position=event.globalPos()
    
    def mover_ventana(self,event):
        if self.isMaximized()==False:
            if event.buttons()==QtCore.Qt.LeftButton:
                self.move(self.pos()+event.globalPos()-self.click_position)
                self.click_position=event.globalPos()
                event.accept()
        if event.globalPos().y()<=10:
            self.showMaximized()
            self.bt_maxi.hide()
            self.bt_restart.show()
        else:
            self.showNormal()
            self.bt_maxi.show()
            self.bt_restart.hide()
                           

 
    def mover_menu(self):
        if True:
            width=self.frame_control.width()
            normal=0
            if width==0:
                extender=200
            else:
                extender=normal
            self.animacion=QPropertyAnimation(self.frame_control,b'minimumWidth')
            self.animacion.setDuration(300)
            self.animacion.setStartValue(width)
            self.animacion.setEndValue(extender)
            self.animacion.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
            self.animacion.start()
    
    def equipo1(self):
        self.bt_load1.setEnabled(True)
     
        
    def load1(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        archivo, _ = QFileDialog.getOpenFileName(self, 'Buscador de csv', '', 'Archivos CSV (*.csv);;Todos los archivos (*)', options=options)
        if archivo:
            self.label_4.setText("Archivo cargado")
            self.archivo_seleccionado = archivo
            self.bt_filter1.setEnabled(True)
        else:
            self.label_4.setText('No se seleccion\xF3 ning\xFAn archivo.')
            self.bt_filter1.setEnabled(False)
    
    def filter1(self):
        if self.archivo_seleccionado:
            try:
                item = self.metodo1.currentText()
                item2 = self.proceso1.currentText()
                self.label_7.setText("Has selecionado\n"+item+" y "+item2)
                df = pd.read_csv(self.archivo_seleccionado)
                columnas = df.columns
                tabla = df[['Abs (Corr)1', 'Abs (Corr)2', 'Abs (Corr)3']]
                self.tabla = tabla
                self.label_4.setText("Archivo Filtrado")
                self.bt_save1.setEnabled(True)
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al filtrar la tabla: {str(e)}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se seleccion\u00F3 ning\xFAn archivo.')
    
    
    def save1(self):
        if hasattr(self, 'tabla'):
            try:
                nombre_archivo = os.path.splitext(os.path.basename(self.archivo_seleccionado))[0]
                nuevo_nombre_excel = f"Filtrado_{nombre_archivo}.xlsx"
                options = QFileDialog.Options()
                options |= QFileDialog.ReadOnly
                ruta_guardado, _ = QFileDialog.getSaveFileName(self, 'Guardar en Excel', nuevo_nombre_excel, 'Archivos Excel (*.xls);;Todos los archivos (*)', options=options)
                
                if ruta_guardado:
                    self.tabla.to_excel(ruta_guardado, index=False)
                    QMessageBox.information(self, 'Informaci\u00F3n', f'Archivo Excel guardado en: {ruta_guardado}')
                    self.label_4.setText("Archivo Guardado")
                    self.bt_save1.setEnabled(False)
                    self.bt_filter1.setEnabled(False)
                    self.label_7.setText('')
                    self.metodo1.setCurrentIndex(0)
                    self.proceso1.setCurrentIndex(0)
                    formato(nuevo_nombre_excel)
                    
                else:
                    QMessageBox.warning(self, 'Advertencia', 'No se selecci\u00F3n una ruta de guardado.')
            except Exception as e:
                QMessageBox.critical(self, 'Error', f'Error al guardar en Excel: {str(e)}')
        else:
            QMessageBox.warning(self, 'Advertencia', 'No se selecci\u00F3n ninguna tabla.')
            
    def formato(name):
        import openpyxl

        # Abre el archivo Excel existente
        archivo_excel = openpyxl.load_workbook(name)

        # Recorre cada hoja en el archivo Excel
        for hoja in archivo_excel.sheetnames:
            hoja_actual = archivo_excel[hoja]

            # Recorre todas las filas y columnas para encontrar la longitud máxima en cada columna
            max_lengths = [0] * (hoja_actual.max_column + 1)

        for fila in hoja_actual.iter_rows(min_row=1, max_row=hoja_actual.max_row, values_only=True):
            for columna, valor in enumerate(fila, start=1):
                if valor is not None:
                    max_lengths[columna] = max(max_lengths[columna], len(str(valor)))

        # Ajusta el ancho de cada columna para acomodar el contenido más largo
        for columna, max_length in enumerate(max_lengths[1:], start=1):
            columna_letra = openpyxl.utils.get_column_letter(columna)
            hoja_actual.column_dimensions[columna_letra].width = max_length + 2  # Agrega un espacio adicional
        
        
        from openpyxl.styles import PatternFill

        # Define el color de fondo amarillo
        relleno_amarillo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

                
        # Aplica el color de fondo amarillo a la primera celda de cada columna
        for columna in hoja_actual.iter_cols(min_col=1, max_col=hoja_actual.max_column, min_row=1, max_row=1):
            for celda in columna:
                celda.fill = relleno_amarillo
        # Guarda los cambios en el archivo Excel
        archivo_excel.save(name)

        # Cierra el archivo Excel
        archivo_excel.close()
    

    def equipo2(self):
        self.bt_load2.setEnabled(True)

    def load2(self):
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getOpenFileName(self, "Select PDF File", "", "PDF files (*.pdf)")
        if file_path:
            document = ap.Document(file_path)
            output_excel, _ = QFileDialog.getSaveFileName(self, "Guardar como archivo Csv", "", "CSV files (*.csv)")

            if output_excel:
               excelSaveOptions = pdf.ExcelSaveOptions()
               excelSaveOptions.format= pdf.ExcelSaveOptions.ExcelFormat.CSV
               document.save(output_excel,excelSaveOptions)
               try:
                   import csv
                   data = []
                   with open(output_excel, newline='') as csvfile:
                        #spamreader = csv.reader(csvfile, delimiter=',') 
                        reader=csv.reader(csvfile)
                        tabla=[]
                        tablas=[]
                        for row in reader:
                            if not row:
                                if tabla:
                                    tablas.append(tabla)
                                    tabla=[]
                            else:
                                tabla.append(row)
                        if tabla:
                            tablas.append(tabla)
                                           
                   for i,tabla in enumerate (tablas):
                       print(f"Tabla{i+1}:") 
                       for fila in tabla:
                           print(', '.join(row))
                   
                   
                
               except Exception as e:
                print("Error al leer el archivo CSV:", str(e))
               
               
               
            # for row in data:
            #     for cell_value in row:
            #         print(cell_value)
                        
                                    
                        
            # Spam, Spam, Spam, Spam, Spam, Baked Beans
            # Spam, Lovely Spam, Wonderful Spam
               # document.save(output_excel, excelSaveOptions)
               # document.save(output_excel, asposepdf.SaveFormat.Excel)
               # df = pd.read_csv(output_excel)
               # columnas = df.columns
               # print(columnas)
                #self.bt_filter2.setEnabled(True)
                

                #for sheet_name in sheet_names:
                    #df = xls.parse(sheet_name, header=None)
                    #self.status_label.setText(f"Mostrando tabla de {sheet_name}")
                    #print(df)
                
                #columnas = df.columns
                #cv_column = df['% CV']
                #tabla = df[['% Triplicados']]
                #self.tabla = cv_column
                #print(cv_column)

                #self.status_label.setText("La conversion se ha completado y las tablas se han mostrado")
        
        
        
            
    def filter2(self, pdf_file):
        self.bt_save2.setEnabled(True)
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        archivo, _ = QFileDialog.getOpenFileName(self, 'Buscador de csv', '', 'Archivos CSV (*.csv);;Todos los archivos (*)', options=options)
        if archivo:
            
            #self.label_4.setText("Archivo cargado")
    
            self.archivo_seleccionado = archivo
            df = pd.read_csv(self.archivo_seleccionado)
            print(df.index)
            
            
        #else:
            #self.label_4.setText('No se seleccion\xF3 ning\xFAn archivo.')
            #self.bt_filter1.setEnabled(False)


    def save2(self):
        if self.file_path:
            document = self.file_path
            doc = ap.Document(self.file_path)
            ruta2, _ = QFileDialog.getSaveFileName(self, "Guardar archivo Excel", "", "Excel files (*.xls)")
            if ruta2:
                save_option = ap.ExcelSaveOptions()
                #self.file_path.to_excel(ruta2, index=False)
                doc.save(ruta2, save_option)
                QMessageBox.information(self, 'Informaci\u00F3n', f'Archivo Excel guardado')
                #save_option.ConvertNonTabularData = True  # Evitar convertir datos no tabulares (CSV)
                #document.save(self.output_excel, save_option)

                #self.status_label.setText("Tablas en formato CSV eliminadas")
                #self.save_button.setEnabled(True)




    


    
    
            
            
          

        
if __name__=="__main__":
    app=QApplication(sys.argv)
    mi_app=rooti()
    mi_app.show()
    sys.exit(app.exec_())